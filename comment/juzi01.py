# @Author:橘子
# @email :2315253816@qq.com
# @Time  :2021/4/21 18:22
# @File  :juzi01.py
import openpyxl
import requests  # 引用第三方库
import time


def read_data(filename, sheetname):
    """
    封装读取函数，定义参数——不需要写死的值
    :param filename: 参数一
    :param sheetname: 参数二
    :return: 返回值
    """
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    case_list = []  # 创建一个列表，用于存放测试用例数据
    max_row = sh.max_row  # 获取到表单的最大行数
    # print("max_row: ", max_row)
    for i in range(2, max_row + 1):  # for循环读取数据表单的每行数据中的url、data、预期结果
        dict1 = dict(  # 把想要获取的数据定义为一个字典，方便后续应用
            case_id=sh.cell(row=i, column=1).value,
            url=sh.cell(row=i, column=5).value,  # 获取这里获取url
            data=sh.cell(row=i, column=6).value,  # 获取请求体
            expect=sh.cell(row=i, column=7).value  # 获取期望结果
        )
        case_list.append(dict1)  # 每循环一次，插入到list末尾

    return case_list  # 定义一个返回值——需要用到的值


def api_fun(url, data):
    """
    定义接口函数参数
    :param url:
    :param data:
    :return:
    """
    header = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
    # 调用post方法返回值
    result = requests.post(url=url, json=data, headers=header).json()
    return result


def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)  # 打开一个已经存在的工作簿
    sh = wb[sheetname]  # 获取表单
    sh.cell(row=row, column=column).value = final_result  # 为第2行第8列数据赋予/修改内容，写入结果
    wb.save(filename)  # 保存文档


def execute_fun(filename, sheetname):
    cases = read_data(filename, sheetname)  # 调用读取表单函数，读取表单
    # print(len(cases))
    # time.sleep(255)
    for case in cases:  # 以次访问cases中的元素
        # print("case: ", case)
        case_id = case['case_id']
        url = case['url']  # 获取url
        data = eval(case['data'])  # 获取data

        expect = eval(case["expect"])
        expect_code = expect['code']  # 获取表单中的期望code
        expect_msg = expect['msg']  # 获取表单中的期望msg
        print("期望结果code为：{}，msg为：{}".format(expect_code, expect_msg))

        real_result = api_fun(url=url, data=data)  # 调用接口测试函数，执行接口测试，把执行结果赋予real_result
        real_code = real_result['code']  # 获取实际运行结果的code
        real_msg = real_result['msg']  # 获取实际运行结果的msg
        print("运行结果code为：{}，msg为：{}".format(real_code, real_msg))

        if real_code == expect_code and real_msg == expect_msg:
            final_re = 'passed'
            print(final_re)
        else:
            final_re = 'failed'
            print(final_re)
        print("*" * 50)
        write_result(filename, sheetname, case_id + 1, 8, final_re)  # 调用写入函数在for遍历下逐条把结果写入第8列


# if __name__ == '__main__':
#     execute_fun('D:\\PycharmProjects\\JK\\test_data\\test_case_api.xlsx', 'login')  # 调用函数

